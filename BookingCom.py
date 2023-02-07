# -*- coding: utf-8 -*-
"""
Created on Sat Jan  7 09:18:12 2023

@author: steven

[金典,福華,裕元,全國,福容	,長榮,清新,林酒店,日月千禧,大毅老爺,日光溫泉,台中艾美,鬱金香,兆品酒店 - 兆尹樓, 兆品酒店 - 品臻樓]
"""
import requests
from bs4 import BeautifulSoup
import pandas as pd
import random
import datetime
import time
from tqdm import tqdm
import openpyxl
from openpyxl.styles import Alignment

def RandomTimeSleep():
    RandomNum = random.uniform(1, 10)
    time.sleep(RandomNum)
    
result=[]
ids = []
data = {}

now = datetime.datetime.now()

hotellist = ['the-splendor-taichung', # 金典
             'prince-hotel-taichung', # 福華
             'windsor-hotel-taichung', # 裕元
             'national-hotel', # 全國
             'fullon-yamay', # 福容
             'evergreen-laurel-of-taichung', # 長榮
             'freshfields-conference-resort', # 清新
             'the-lin-hotel', # 林酒店
             'millennium-vee-taichung', # 日月千禧
             'tai-zhong-da-yi-lao-ye-xing-lu', # 大毅老爺
             'the-sun-hot-spring-resort', # 日光溫泉
             'le-meridien-taichung', # 台中艾美
             'tai-zhong-zhen-da-jin-yu-jin-xiang-jiu-dian', # 震大金鬱金香酒店
             'maison-de-chine-taichung', # 兆品酒店 - 兆尹樓
             'maison-de-chine-taichung-pin-chen-building' # 兆品酒店 - 品臻樓
             ]

checkin = now.strftime("%Y-%m-%d")
checkout = (now + datetime.timedelta(days=1)).strftime("%Y-%m-%d")

#excel_filename = 'BookingCom'+checkin+' - '+checkout+'.xlsx'
excel_filename = './OTAExcel/BookingCom.xlsx'

for hotelnameen in tqdm(hotellist):
    RandomTimeSleep()
    
    url = "https://www.booking.com/hotel/tw/"+hotelnameen+".zh-tw.html?aid=304142&label=gen173nr-1FCAEoggI46AdIMFgEaOcBiAEBmAEwuAEXyAEM2AEB6AEB-AECiAIBqAIDuAKcyd2dBsACAdICJDA1MmE4YzFmLWQzNjEtNDRkZS04MWQxLTVmMmI0MjZmNzc5Y9gCBeACAQ&sid=7c8766b4cd3a131788f69ea01ad4f3b2&all_sr_blocks=31256504_91600219_0_1_0;checkin="+checkin+";checkout="+checkout+";dest_id=312565;dest_type=hotel;dist=0;group_adults=1;group_children=0;hapos=1;highlighted_blocks=31256504_91600219_0_1_0;hpos=1;matching_block_id=31256504_91600219_0_1_0;no_rooms=1;req_adults=1;req_children=0;room1=A;sb_price_type=total;sr_order=popularity;sr_pri_blocks=31256504_91600219_0_1_0__520000;srepoch=1672964120;srpvid=8c3701cbd1a30025;type=total;ucfs=1&#hotelTmpl"
    
    headers={"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36"}
    
    resp = requests.get(url, headers=headers)
    
    if resp.ok:
        soup = BeautifulSoup(resp.text, 'html.parser')
        
        HotelName = soup.find("h2",{"class":"pp-header__title"}).text
        
        try:
            tr = soup.find_all("tr")
        except:
            tr = None
        
        for y in range(0,len(tr)):
            try:
                id = tr[y].get('data-block-id')
            except:
                id = None
                
            if( id is not None):
                ids.append(id)
        
        for i in range(0,len(ids)):
            try:
                allData = soup.find("tr",{"data-block-id":ids[i]})
                try:
                    rooms = allData.find("span",{"class":"hprt-roomtype-icon-link"})
                except:
                    rooms = None
                if(rooms is not None):
                    last_room = rooms.text.replace("\n","").strip()
                try:
                    data["RoomName"] = rooms.text.replace("\n","").strip()
                except:
                    data["RoomName"] = last_room
                
                original = allData.find("div",{"class":"bui-f-color-destructive js-strikethrough-price prco-inline-block-maker-helper bui-price-display__original"})
                try:
                    data["Original"] = original.text.replace("\n","").strip().strip('TWD\xa0').replace(',','')
                except:
                    data["Original"] = original 
                
                price = allData.find("div",{"class":"bui-price-display__value prco-text-nowrap-helper prco-inline-block-maker-helper prco-f-font-heading"})
                data["Price"] = price.text.replace("\n","").strip().strip('TWD\xa0').replace(',','')

                
                data["HotelName"] = HotelName
        
                result.append(data)
                data={}
            except:
                data["RoomName"] = None
                data["Price"] = None

        for i in range(0,len(soup.find_all("div",{"class":"d46673fe81 f16339d4be cbc2fe2dfe c6aefe00bc c135d5bf2d"}))):
            data["HotelName"] = HotelName
            data["RoomName"] = soup.find_all("a",{"class":"fc63351294 a168c6f285 d1c4779e7a js-legacy-room-name a25b1d9e47"})[i].text.replace("\n","").strip()
            data["Original"] = ''
            data["Price"] = soup.find_all("div",{"class":"db29ecfbe2 b028a54d7f"})[i].text.replace("\n","").strip()

            result.append(data)
            data={}
            
df = pd.DataFrame(result, columns=['HotelName', 'RoomName', 'Original', 'Price'])
df = df.rename(columns={'HotelName': '飯店名稱', 'RoomName': '房型', 'Original': '原價', 'Price': '售價'})
df.head()
df.to_excel(excel_filename, index=None, startrow = 1, sheet_name = "Booking.com")

wb = openpyxl.load_workbook(excel_filename)
ws = wb.active
ws['A1'].value = '訂房日期'+checkin+' - '+checkout
align = Alignment(horizontal='center', vertical='center',wrap_text=True)

ws['A1'].alignment = align

ws.merge_cells('A1:D1')

wb.save(excel_filename)